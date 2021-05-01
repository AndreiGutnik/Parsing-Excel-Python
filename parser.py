import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Font, Side

# ---- ВЫБОР ВСЕХ ФАЙЛОВ EXCEL ----

PATH = 'D:\Python\Parsing Excel Python\Parsing Excel Python'

files = []
def obchodfiles(path, level=1):
    for i in os.listdir(path):
        if os.path.isdir(path + '\\' + i):
            obchodfiles(path + '\\' + i, level=level + 1)
        else:
            # print(PATH + '\\' + i)
            if i.endswith('.xlsx'):
                files.append(i)

obchodfiles(PATH)
print(files)

# ----ЧТЕНИЕ ПОЛУЧЕННЫХ ФАЙЛОВ ----

FILE = 'Test.xlsx'

# wb = load_workbook(FILE) -- Получение формул
wb = load_workbook(FILE, data_only=True)  # Получение значений

ws = wb.create_sheet("Mysheet")  # Создание нового листа
print(wb.sheetnames[3])
wb.remove(ws)  # Удаление листа
print(wb.sheetnames)  # Список всех листов

# ДОСТУП К ЯЧЕЙКАМ
print(wb['Лист 3']['C3'].value)

wb.active = 1
sheet = wb.active  # wb['Лист 2']
print(sheet['B3'].value)

for i in wb.sheetnames:
    print(wb[i]['A2'].value)

print(sheet.max_row)  # Количество строк
print(sheet.max_column)  # Количество столбцов

print('\n--- Чтение файла ---\n')
# end = sheet.cell(sheet.max_row, sheet.max_column).coordinate
# data = sheet['A1':end]

shapka = [cell.value for cell in next(sheet.iter_rows(min_row=2, min_col=1, max_row=2,
                                                      max_col=sheet.max_column))]
# print(shapka)

data = {}
usl_data = []
for row in sheet.iter_rows(min_row=3, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
    if len(row) > 0:
        usl = row[1].value  # Чтение столбца
        if usl is not None:
            usl_data = [cell.value for cell in row]  # Непосредственное чтение строк
            if usl not in data:  # Создание массива данных по услуге
                data[usl] = []
            data[usl].append(usl_data)

wb.close()

# Создание отчетов по каждой услуге и сохранение в файл
for usl in data:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Услуги'

    ws.append(shapka)
    for row in data[usl]:
        ws.append(row)

# Форматирование шапки
    for i in range(1,7):
        zagl = ws.cell(row=1, column=i)
        zagl.alignment = Alignment(horizontal='center', vertical='center')
        zagl.fill = PatternFill(fill_type='solid', start_color='5a61f0', end_color='5a61f0')
        zagl.font = Font(bold=True, italic=True, color='ffffff', size='16')
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="ff0000")
        zagl.border = Border(top=double, left=thin, right=thin, bottom=double)

# Форматирование содержимого по столбцам
    nmrow = len(data[usl])
    for i in range(2, nmrow+2):
        ws.cell(row=i, column=1).number_format = '# ##0'
        ws.cell(row=i, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=i, column=1).fill = PatternFill(fill_type='solid', start_color='f8ca30', end_color='f8ca30')

        ws.cell(row=i, column=2).alignment = Alignment(horizontal='left')
        ws.cell(row=i, column=2).fill = PatternFill(fill_type='solid', start_color='f8ca30', end_color='f8ca30')


# Ширина столбцов
    ws.row_dimensions[1].height = 30
    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 12

    exfilname = os.path.join('.', 'Data', (usl + '.xlsx'))
    exfilname = os.path.abspath(exfilname)
    print(exfilname)
    wb.save(exfilname)
    wb.close()